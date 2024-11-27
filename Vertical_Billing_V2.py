# Make changes in this if required
# mail.To = "corazon.malanes@sutherlandglobal.com; rajesh.a@sutherlandglobal.com; Ganesh.Pillai@SutherlandGlobal.COM; Roel.Florendo@sutherlandglobal.COM; VincelMark.Costa@SUTHERLANDGLOBAL.COM; nitin.rao@sutherlandglobal.com; abhishek.raj@sutherlandglobal.com;businessfinancetechnology@sutherlandglobal.com"

# mail.CC = "GlobalCapPlanTeam@sutherlandglobal.com; Mohammed.Zeeshan1@sutherlandglobal.com"

import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import win32com.client as win32

# Get the current working directory
notebook_directory = os.path.abspath('')

# Define the source Excel file path
source_excel_path = os.path.join(notebook_directory, "Consolidated billing.xlsx")

# Read the source Excel file into a DataFrame
df = pd.read_excel(source_excel_path)

# Filter the DataFrame based on the "Vertical" column
filtered_df = df[df["Vertical"] == "Technology_v"]

# Get the current month and year for the new file name
current_date = datetime.now()
new_file_name = f"Consolidated_Billing_Inputs_Technology_Vertical_{current_date.strftime('%B')}'{current_date.strftime('%y')}.xlsx"

# Define the path for the new Excel file
new_excel_path = os.path.join(notebook_directory, new_file_name)

# Write the filtered data to the new Excel file
filtered_df.to_excel(new_excel_path, sheet_name="Sheet1", index=False)

# Load the saved workbook using openpyxl
saved_workbook = openpyxl.load_workbook(new_excel_path)
saved_worksheet = saved_workbook.active

# Define a fill pattern for light green background
light_green_fill = PatternFill(start_color='90ee90', end_color='90ee90', fill_type='solid')

# Apply the light green background to the first row (header)
for cell in saved_worksheet[1]:
    cell.fill = light_green_fill

# Reduce the font size of the rest of the rows by 10%
font = Font(size=10)
for row in saved_worksheet.iter_rows(min_row=2, values_only=False):
    for cell in row:
        cell.font = font

# Set column width to auto-adjust based on content
for column in saved_worksheet.columns:
    max_length = max(len(str(cell.value)) for cell in column if cell.value)
    adjusted_width = max_length + 2
    saved_worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook with the updated formatting and column widths
saved_workbook.save(new_excel_path)

# Initialize Outlook
outlook = win32.Dispatch('Outlook.Application')

# Create a new email
mail = outlook.CreateItem(0)
mail.Subject = current_date.strftime("Technology Vertical Billing Inputs - %B %Y")
mail.Body = f'''Hi All,
\n\nPlease find attached the billing inputs for the month of {current_date.strftime('%B %Y')}, for technology vertical. Request you to check and confirm on the same.
\n\nRequest you to pass on this mail to respective stakeholders in case I have missed any.
\n\nRegards,
\n\nMohammed Zeeshan
\nLead- Capacity Planner
\nPh: +91 988429140'''
mail.To = "mohammed.zeeshan1@sutherlandglobal.com"
mail.CC = "mohammed.zeeshan1@sutherlandglobal.com; Mohammed.Zeeshan1@sutherlandglobal.com"

# Attach the Excel file
mail.Attachments.Add(new_excel_path)

# Send the email
mail.Send()
print("Mail Sent")