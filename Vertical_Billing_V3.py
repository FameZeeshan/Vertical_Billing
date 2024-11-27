import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import win32com.client as win32
import time

# Get the current working directory
notebook_directory = os.path.abspath('')

# Define the source Excel file path
source_excel_path = os.path.join(notebook_directory, "Consolidated billing.xlsx")

# Read the source Excel file into a DataFrame
df = pd.read_excel(source_excel_path)

# Get unique values from the "Vertical" column
unique_verticals = df["Vertical"].unique()

# Get the current month and year for the new file name
current_date = datetime.now()

# Define mailto lists for each vertical (example lists)
mailto_lists = {
    "Technology_v": ["Mohammed.Zeeshan1@sutherlandglobal.com", "Mohammed.Zeeshan1@sutherlandglobal.com"],
    "Amazon_v": ["Mohammed.Zeeshan1@sutherlandglobal.com", "Mohammed.Zeeshan1@sutherlandglobal.com"],
    "CME_v": ["Mohammed.Zeeshan1@sutherlandglobal.com", "Mohammed.Zeeshan1@sutherlandglobal.com"]
}

# Define a common CC mail list
cc_mail_list = ["Mohammed.Zeeshan1@sutherlandglobal.com", "Mohammed.Zeeshan1@sutherlandglobal.com"]

# Initialize Outlook
outlook = win32.Dispatch('Outlook.Application')

# Loop through each unique vertical and send emails accordingly
for vertical in unique_verticals:
    try:
        # Filter the DataFrame for the current vertical
        filtered_df = df[df["Vertical"] == vertical]
        
        # Create a new file name for the current vertical
        new_file_name = f"Consolidated_Billing_Inputs_{vertical}_{current_date.strftime('%B')}'{current_date.strftime('%y')}.xlsx"
        
        # Define the path for the new Excel file
        new_excel_path = os.path.join(notebook_directory, new_file_name)
        
        # Write the filtered data to the new Excel file
        with pd.ExcelWriter(new_excel_path, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, sheet_name=vertical, index=False)
        
        # Load the saved workbook using openpyxl
        saved_workbook = openpyxl.load_workbook(new_excel_path)
        saved_worksheet = saved_workbook.active
        
        # Define a fill pattern for light green background
        light_green_fill = PatternFill(start_color='90ee90', end_color='90ee90', fill_type='solid')
        
        # Apply the light green background to the first row (header)
        for cell in saved_worksheet[1]:
            cell.fill = light_green_fill
        
        # Reduce the font size of the rest of the rows to 10
        font = Font(size=10)
        for row in saved_worksheet.iter_rows(min_row=2, values_only=False):
            for cell in row:
                cell.font = font
        
        # Set column width to auto-adjust based on content
        for column in saved_worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = max_length + 2
            saved_worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Save the workbook with the updated formatting and column widths
        saved_workbook.save(new_excel_path)
        
        # Check if the vertical has a mailto list
        if vertical in mailto_lists and mailto_lists[vertical]:
            # Create a new email
            mail = outlook.CreateItem(0)
            mail.Subject = current_date.strftime(f"{vertical} Vertical Billing Inputs - %B %Y")
            mail.Body = f'''Hi All,
            Please find attached the billing inputs for the month of {current_date.strftime('%B %Y')}, for technology vertical. Request you to check and confirm on the same.
            Request you to pass on this mail to respective stakeholders in case I have missed any.
            \nRegards,
            Mohammed Zeeshan
            Lead- Capacity Planner
            Ph: +91 988429140'''
            # Set recipients based on vertical-specific mailto list
            mail.To = "; ".join(mailto_lists[vertical])
            
            # Set CC recipients from common CC mail list
            mail.CC = "; ".join(cc_mail_list)
            
            # Attach the Excel file containing only the current vertical sheet
            mail.Attachments.Add(new_excel_path)
            
            # Send the email
            mail.Send()
            print(f"Mail sent for vertical: {vertical}")
            
            # Add a delay of 1 second before sending the next email
            time.sleep(1)
            
        else:
            print(f"No mailto list found for vertical: {vertical}")
    except Exception as e:
        print(f"Failed to process vertical: {vertical}. Error: {e}")

print("Mails Sent")